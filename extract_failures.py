import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read the categories.json file
with open('/mnt/e/chromedoanload/allure-report/allure-report/data/categories.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Extract failed test cases
failed_cases = []

# Process Product defects category
for category in data.get('children', []):
    category_name = category.get('name', '')
    for defect in category.get('children', []):
        defect_name = defect.get('name', '')
        
        # Extract error message from defect name
        error_msg = defect_name
        
        for test_case in defect.get('children', []):
            test_name = test_case.get('name', '')
            uid = test_case.get('uid', '')
            status = test_case.get('status', '')
            parameters = test_case.get('parameters', [])
            
            # Extract metric name, function, dimensions from parameters
            metric_name = parameters[0] if len(parameters) > 0 else ''
            function = parameters[1] if len(parameters) > 1 else ''
            dimensions = parameters[4] if len(parameters) > 4 else ''
            
            failed_cases.append({
                'uid': uid,
                'test_name': test_name,
                'metric_name': metric_name,
                'function': function,
                'dimensions': dimensions,
                'error_message': error_msg,
                'category': category_name,
                'status': status
            })

print(f"Total failed cases: {len(failed_cases)}")

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = "失败用例清单"

# Header styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write headers
headers = ['序号', 'UID', '测试名称', '指标名称', '函数', '维度', '错误信息', '分类', '状态']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Write data
for row, case in enumerate(failed_cases, 2):
    ws.cell(row=row, column=1, value=row-1)
    ws.cell(row=row, column=2, value=case['uid'])
    ws.cell(row=row, column=3, value=case['test_name'])
    ws.cell(row=row, column=4, value=case['metric_name'])
    ws.cell(row=row, column=5, value=case['function'])
    ws.cell(row=row, column=6, value=case['dimensions'])
    ws.cell(row=row, column=7, value=case['error_message'])
    ws.cell(row=row, column=8, value=case['category'])
    ws.cell(row=row, column=9, value=case['status'])

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 60
ws.column_dimensions['G'].width = 80
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 10

# Save
output_path = '/mnt/f/.openclaw/workspace/失败用例清单.xlsx'
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
print(f"Total rows: {len(failed_cases)}")
